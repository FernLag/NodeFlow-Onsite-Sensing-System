"""
build.py - regenerates the config blocks in assets/js/main.js from
data/sensor_configuration.xlsx

Run it from the repository root:  python3 tools/build.py

This script rewrites these blocks in assets/js/main.js based on the Excel file:
  - SENSOR_TYPES     ← sheet "sensors" + "params"
  - OUTPUT_PARAMS    ← sheet "sensors" (in_a..in_e columns)
  - OUTPUT_VIZ       ← sheet "sensors" (new "viz" column, pipe-separated viz_keys)
  - PORT_TIPS, PORTS ← sheet "ports"
  - VIZ_OPTIONS      ← sheet "viz_options"
  - SURVEY_QUESTIONS ← sheet "survey_questions"

It does NOT touch the TEMPLATES block - that contains the actual Arduino
code logic and must be edited by hand in assets/js/main.js.

IMPORTANT: For Excel edits to flow through to generated code, the output names
in your "sensors" sheet (e.g. "TAW", "Raw value") must match keys in
TEMPLATES.outputs (case-insensitive). If you add a new output in Excel, add a
matching entry in TEMPLATES.outputs in assets/js/main.js.

Same rule applies for new sensor types (TEMPLATES.sensors) and new viz
options (TEMPLATES.viz).
"""

import os
import re
import sys
from openpyxl import load_workbook

# Paths are resolved against the repository root so the script can be run from
# anywhere:  python3 tools/build.py
ROOT          = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_NAME    = "sensor_configuration.xlsx"
MAIN_JS_FILE  = os.path.join(ROOT, "assets", "js", "main.js")
TEMPLATES_DIR = os.path.join(ROOT, "tools", "templates")


def find_excel():
    """Prefer the edited copy if there is one, so the original master is never
    written to and can always be diffed against. Older checkouts kept the file
    at the repository root, so fall back to that before giving up."""
    for candidate in (os.path.join(ROOT, "data", "sensor_configuration_v2.xlsx"),
                      os.path.join(ROOT, "data", EXCEL_NAME),
                      os.path.join(ROOT, EXCEL_NAME)):
        if os.path.isfile(candidate):
            return candidate
    sys.exit(
        f"Could not find {EXCEL_NAME}.\n"
        f"Put the master spreadsheet in {os.path.join(ROOT, 'data')} and run this again."
    )


EXCEL_FILE = find_excel()
wb = load_workbook(EXCEL_FILE, data_only=True)
print(f"reading {os.path.relpath(EXCEL_FILE, ROOT)}")

def sheet_rows(sheet_name, skip=1):
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=skip + 1, values_only=True):
        if any(c is not None for c in row):
            rows.append([str(c).strip() if c is not None else "" for c in row])
    return rows

def safe_js_key(key):
    return re.sub(r'[^a-zA-Z0-9_]', '_', key)



SENSOR_CANON = {
    "df_robot":              "DF_robot",
    "dfrobot":               "DF_robot",
    "watermark":             "Watermark",
    "watermark_temperature": "Watermark_Temperature",
}

def canonical_sensor(name):
    n = (name or "").strip().lower()
    if n in SENSOR_CANON:
        return SENSOR_CANON[n]
    if "capacitive" in n or "dfrobot" in n:
        return "DF_robot"
    
    if ("temp" in n or "200ts" in n) and "watermark" not in n and "combined" not in n:
        return "Temperature"
    
    if "combined" in n or "temp" in n:
        return "Watermark_Temperature"
    if "watermark" in n or "irrometer" in n:
        return "Watermark"
    return safe_js_key(name)

def sheet_header(sheet_name):
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c).strip().lower() if c is not None else "" for c in row]
    return []

sensors_header = sheet_header("sensors")
def col_index(name, default):
    try:
        return sensors_header.index(name.lower())
    except ValueError:
        return default

VIZ_COL = col_index("viz", 11)

sensors_raw = sheet_rows("sensors")


COL = {
    "sensor_name":  col_index("sensor_name", 0),
    "output_value": col_index("output_value", 1),
    "full_name":    col_index("full_name", 2),
    "info_tip":     col_index("info_tip", 3),
    "in_a":         col_index("in_a", 4),
    "in_b":         col_index("in_b", 5),
    "in_c":         col_index("in_c", 6),
    "in_d":         col_index("in_d", 7),
    "in_e":         col_index("in_e", 8),
    "equation":     col_index("equation", 9),
    "viz":          col_index("viz", 11),
}

_expected = {"sensor_name", "output_value", "full_name", "info_tip",
             "in_a", "in_b", "in_c", "in_d", "in_e", "equation", "viz"}
_found = {h for h in sensors_header if h in _expected}
if _expected - _found:
    print("  ! sensors sheet is missing header(s):", ", ".join(sorted(_expected - _found)))
    print("    falling back to default column positions for those.")

def cell(row, name):
    i = COL[name]
    return row[i] if i is not None and len(row) > i else ""

UNFINISHED = re.compile(r"not yet implemented", re.I)

active_keys = []
sensor_meta = {}
outputs_map = {}
skipped_outputs = []

current_key = ""
for row in sensors_raw:
    key      = cell(row, "sensor_name")
    out_val  = cell(row, "output_value")
    out_full = cell(row, "full_name")
    out_tip  = cell(row, "info_tip")
    inp_a    = cell(row, "in_a")
    inp_b    = cell(row, "in_b")
    inp_c    = cell(row, "in_c")
    inp_d    = cell(row, "in_d")
    inp_e    = cell(row, "in_e")
    equation = cell(row, "equation")
    viz_list = cell(row, "viz")

    if key:
        current_key = canonical_sensor(key)
        if current_key not in sensor_meta:
            sensor_meta[current_key] = {
                "label":   key,
                "tip":     out_tip,
                "raw_key": key,
            }
            active_keys.append(current_key)

    if not current_key:
        continue

    # A measurement whose own tooltip says it is unfinished should not be on
    # offer: it appears in the dropdown and generates a plausible looking
    # number, which is worse than not being there. The sheet already states
    # the status in prose, so that is what is honoured here rather than a
    # second list to keep in step. Take the marker out of the tooltip and the
    # measurement comes back.
    if out_val and UNFINISHED.search(out_tip or ""):
        skipped_outputs.append(f"{current_key}: {out_val}")
        continue

    if out_val and out_val.lower() != "for all params, col min max default":
        outputs_map.setdefault(current_key, [])
        if not any(o["value"] == out_val for o in outputs_map[current_key]):
            outputs_map[current_key].append({
                "value":    out_val,
                "tip":      out_tip,
                "full":     out_full,
                "equation": equation,
                "inputs":   [inp_a, inp_b, inp_c, inp_d, inp_e],
                "viz":      viz_list,
            })

params_raw = sheet_rows("params")
params_map = {}

for row in params_raw:
    key       = canonical_sensor(row[0])
    p_name    = row[1]
    
    if len(row) >= 8 and row[2] and not str(row[2]).replace(".", "").replace("-", "").lstrip().isdigit():
        p_display = row[2]
        p_label   = row[3]
        p_min     = row[4]
        p_max     = row[5]
        p_val     = row[6] if len(row) > 6 else "0"
        p_units   = row[7] if len(row) > 7 else ""
        p_choices = row[8] if len(row) > 8 else ""
    else:
        p_display = ""
        p_label   = row[2]
        p_min     = row[3]
        p_max     = row[4]
        p_val     = row[5] if len(row) > 5 else "0"
        p_units   = row[6] if len(row) > 6 else ""
        p_choices = row[7] if len(row) > 7 else ""

    if not key or not p_name:
        continue

    params_map.setdefault(key, [])
    params_map[key].append({
        "name":    p_name,
        "display": p_display if p_display else p_name,
        "label":   p_label,
        "min":     p_min,
        "max":     p_max,
        "value":   str(p_val) if p_val != "" else "0",
        "units":   str(p_units) if p_units else "",
        "choices": str(p_choices) if p_choices else "",
    })

viz_raw      = sheet_rows("viz_options")
active_viz   = [r for r in viz_raw if r[3].upper() == "TRUE"]
survey_raw   = sheet_rows("survey_questions")
ports_raw    = sheet_rows("ports")
active_ports = [r[0] for r in ports_raw if r[0]]
port_tips    = {r[0]: r[1] for r in ports_raw if r[0]}


def jstr(s):
    return (s or "").replace("\\", "\\\\").replace('"', '\\"')


def build_sensor_types():
    lines = ["const SENSOR_TYPES = {"]
    for key in active_keys:
        meta    = sensor_meta[key]
        outputs = outputs_map.get(key, [])
        params  = params_map.get(key, [])
        lines.append(f'  {key}: {{')
        lines.append(f'    label: "{jstr(meta["label"])}",')
        lines.append(f'    tip: "{jstr(meta["tip"])}",')
        lines.append('    outputs: [')
        for o in outputs:
            lines.append(f'      {{')
            lines.append(f'        value: "{jstr(o["value"])}",')
            lines.append(f'        display: "{jstr(o.get("full") or o["value"])}",')
            lines.append(f'        tip: "{jstr(o["tip"])}",')
            lines.append(f'      }},')
        lines.append('    ],')
        lines.append('    params: [')
        for p in params:
            lines.append(f'      {{')
            lines.append(f'        name: "{jstr(p["name"])}",')
            lines.append(f'        display: "{jstr(p["display"])}",')
            lines.append(f'        label: "{jstr(p["label"])}",')
            lines.append(f'        value: "{jstr(p["value"])}",')
            lines.append(f'        min: "{jstr(p["min"])}",')
            lines.append(f'        max: "{jstr(p["max"])}",')
            lines.append(f'        units: "{jstr(p["units"])}",')
            if p.get("choices"):
                lines.append(f'        choices: "{jstr(p["choices"])}",')
            lines.append(f'      }},')
        lines.append('    ],')
        lines.append('  },')
    lines.append('};')
    return '\n'.join(lines)


def build_port_tips():
    lines = ["const PORT_TIPS = {"]
    for port, tip in port_tips.items():
        lines.append(f'  {port}: "{jstr(tip)}",')
    lines.append('};')
    return '\n'.join(lines)


def build_ports():
    items = ',\n'.join(f'  "{p}"' for p in active_ports)
    return f'const PORTS = [\n{items},\n];'


def build_viz_options():
    lines = ["const VIZ_OPTIONS = ["]
    for row in active_viz:
        key, label, tip = row[0], row[1], row[2]
        lines.append(f'  {{')
        lines.append(f'    value: "{jstr(key)}",')
        lines.append(f'    label: "{jstr(label)}",')
        lines.append(f'    tip: "{jstr(tip)}",')
        lines.append(f'  }},')
    lines.append('];')
    return '\n'.join(lines)


def build_output_params():
    lines = ["const OUTPUT_PARAMS = {"]
    for key in active_keys:
        outputs = outputs_map.get(key, [])
        lines.append(f'  {key}: {{')
        for o in outputs:
            inputs = [str(i).strip() for i in o["inputs"] if i not in (None, "", " ")]
            arr = "[" + ", ".join(f'"{jstr(i)}"' for i in inputs) + "]"
            lines.append(f'    "{jstr(o["value"])}": {arr},')
        lines.append('  },')
    lines.append('};')
    return '\n'.join(lines)


def build_output_viz():
    lines = ["const OUTPUT_VIZ = {"]
    for key in active_keys:
        outputs = outputs_map.get(key, [])
        lines.append(f'  {key}: {{')
        for o in outputs:
            raw = o.get("viz", "") or ""
            vizzes = [v.strip() for v in str(raw).split("|") if v.strip()]
            if not vizzes:
                vizzes = ["none"]
            if "none" not in vizzes:
                vizzes = ["none"] + vizzes
            arr = "[" + ", ".join(f'"{jstr(v)}"' for v in vizzes) + "]"
            lines.append(f'    "{jstr(o["value"])}": {arr},')
        lines.append('  },')
    lines.append('};')
    return '\n'.join(lines)


def build_survey_questions():
    lines = ["const SURVEY_QUESTIONS = ["]
    for row in survey_raw:
        key, label, qtype = row[0], row[1], row[2]
        required = "true" if row[3].upper() == "TRUE" else "false"
        extra    = row[4] if len(row) > 4 else ""
        lines.append(f'  {{')
        lines.append(f'    key: "{key}",')
        lines.append(f'    label: "{jstr(label)}",')
        lines.append(f'    type: "{qtype}",')
        lines.append(f'    required: {required},')
        if qtype.lower() == "select":
            options = [o.strip() for o in extra.split("|") if o.strip()]
            opts_js = "[" + ", ".join(f'"{jstr(o)}"' for o in options) + "]"
            lines.append(f'    options: {opts_js},')
        else:
            lines.append(f'    placeholder: "{jstr(extra)}",')
        lines.append(f'  }},')
    lines.append('];')
    return '\n'.join(lines)


with open(MAIN_JS_FILE, "r", encoding="utf-8") as f:
    js = f.read()

js = re.sub(r'const SENSOR_TYPES = \{.*?\};',
            build_sensor_types(), js, flags=re.DOTALL)
js = re.sub(r'const PORT_TIPS = \{.*?\};',
            build_port_tips(), js, flags=re.DOTALL)
js = re.sub(r'const PORTS = \[.*?\];',
            build_ports(), js, flags=re.DOTALL)
js = re.sub(r'const VIZ_OPTIONS = \[.*?\];',
            build_viz_options(), js, flags=re.DOTALL)
js = re.sub(r'const SURVEY_QUESTIONS = \[.*?\n\];',
            build_survey_questions(), js, flags=re.DOTALL)
js = re.sub(r'const OUTPUT_PARAMS = \{.*?\n\};',
            build_output_params(), js, flags=re.DOTALL)
js = re.sub(r'const OUTPUT_VIZ = \{.*?\n\};',
            build_output_viz(), js, flags=re.DOTALL)

with open(MAIN_JS_FILE, "w", encoding="utf-8") as f:
    f.write(js)

if skipped_outputs:
    print("Left out, because the tooltip says the measurement is not yet implemented:")
    for entry in skipped_outputs:
        print("   ", entry)

print(f"assets/js/main.js updated: {len(active_keys)} sensor(s), {len(active_ports)} port(s), "
      f"{len(active_viz)} viz option(s), {len(survey_raw)} survey question(s), "
      f"OUTPUT_PARAMS + OUTPUT_VIZ regenerated")


os.makedirs(TEMPLATES_DIR, exist_ok=True)

for key in active_keys:
    meta    = sensor_meta[key]
    params  = params_map.get(key, [])
    outputs = outputs_map.get(key, [])

    param_lines = []
    for p in params:
        mn = f"  # min: {p['min']}" if p["min"] else ""
        mx = f"  max: {p['max']}" if p["max"] else ""
        param_lines.append(f'{p["name"]} = {p["value"]}  # {p["label"]}{mn}{mx}')
    param_block = "\n".join(param_lines) if param_lines else "# No parameters defined"

    output_lines = []
    for o in outputs:
        eq     = f"  # equation: {o['equation']}" if o["equation"] else ""
        inputs = [i for i in o["inputs"] if i]
        inp    = f"  # inputs: {', '.join(inputs)}" if inputs else ""
        output_lines.append(f'#   - {o["value"]}{inp}{eq}')
    output_block = "\n".join(output_lines) if output_lines else "#   (none defined)"

    p0 = params[0]["name"] if len(params) > 0 else "param_1"
    p1 = params[1]["name"] if len(params) > 1 else "param_2"

    code = f'''# {key}.py
# Sensor: {meta["label"]}

{param_block}

# Supported outputs:
{output_block}


def read(raw_value: float) -> dict:
    result = {{}}
    return result


if __name__ == "__main__":
    test_raw = 512
    output   = read(test_raw)
    print(f"Sensor : {meta['label']}")
    print(f"Raw    : {{test_raw}}")
    print(f"Output : {{output}}")
'''

    path = os.path.join(TEMPLATES_DIR, f"{key}.py")
    with open(path, "w", encoding="utf-8") as f:
        f.write(code)
    print(f"tools/templates/{key}.py written")

print("Done.")